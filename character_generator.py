# Note: This program requires the openpyxl module.
# Install it with: pip install openpyxl
# ========================
# PART 1: IMPORTS & CONSTANTS 
# ========================

import random
import openpyxl

# ========================
# PART 2: WORLD BACKGROUND & REGIONS
# ========================

# Define the four regions with their themes, allowed races, and major cities

REGIONS = {
    "Borealans": {
        "theme": "Frozen tundra / taiga, harsh winters, long days (northern / Nordic)",
        "races": ["Human", "Orc", "Dwarf"],
        "cities": ["Hinterborne", "Frostmoor", "Drakkenstone"]
    },
    "Tropicanis": {
        "theme": "Southern tropics, rainforest, very humid, lots of life",
        "races": ["Human", "Saurian", "Halfling"],
        "cities": ["Everspring", "Doradel", "Calthos", "Langrisha", "Hambshala", "Lavayon"]
    },
    "Selvarossa": {
        "theme": "Island culture (Oceania-like)",
        "races": ["Human", "Aelf", "Faerie"],
        "cities": ["Bellariva", "Castrovivo", "Vetricella", "Castanera", "Rovella", "Forcillia"]
    },
    "Verdanian": {
        "theme": "Frontier world (colonial frontier America feel)",
        "races": ["Centaur", "Dwarf", "Aelf", "Orc"],
        "cities": ["Thistleton", "Skylorne", "Thornfell", "Redhook"]
    }
}

# List of possible genders
GENDERS = ["Male", "Female", "Nonbinary"]


# Randomly choose a region, race, city, and gender for the character
def generate_background():
    # pick a random region
    region_name = random.choice(list(REGIONS.keys()))
    region_info = REGIONS[region_name]
    
    # pick a race from that region
    race = random.choice(region_info["races"])
    
    # pick a city
    city = random.choice(region_info["cities"])
    
    # pick a gender
    gender = random.choice(GENDERS)
    
    # put it all together
    background = {
        "region": region_name,
        "city": city,
        "race": race,
        "gender": gender
    }
    
    return background


# ========================
# PART 3: PHYSICAL CHARACTERISTICS
# ========================

# Define height ranges for each race (stored as min and max in inches)
RACE_HEIGHT_RANGES = {
    "Human": (58, 78),      # 4'10" to 6'6"
    "Orc": (70, 90),        # 5'10" to 7'6"
    "Dwarf": (48, 65),      # 4'0" to 5'5"
    "Saurian": (54, 72),    # 4'6" to 6'0"
    "Halfling": (42, 58),   # 3'6" to 4'10"
    "Aelf": (67, 82),       # 5'7" to 6'10"
    "Faerie": (36, 54),     # 3'0" to 4'6"
    "Centaur": (72, 108)    # 6'0" to 9'0"
}

# Define weight ranges for each race (stored as min and max in pounds)
RACE_WEIGHT_RANGES = {
    "Human": (80, 400),
    "Orc": (250, 600),
    "Dwarf": (200, 400),
    "Saurian": (150, 300),
    "Halfling": (40, 150),
    "Aelf": (120, 200),
    "Faerie": (30, 100),
    "Centaur": (1000, 2500)
}

# Define hair color options for each race
RACE_HAIR_COLORS = {
    "Human": ["Black", "Brown", "Red", "Blonde", "Gray"],
    "Orc": ["Black"],
    "Dwarf": ["Black", "Brown", "Red", "Blonde", "Gray"],
    "Saurian": ["None"],  # Saurians don't have hair
    "Halfling": ["Black", "Brown", "Red", "Blonde", "Gray"],
    "Aelf": ["Blonde", "White", "Silver", "Copper", "Black"],
    "Faerie": ["Black", "Brown", "Red", "Orange", "Yellow", "Green", "Blue", "Purple"],
    "Centaur": ["Brown", "White", "Black", "Gray", "Auburn"]
}

# Define eye color options for each race
RACE_EYE_COLORS = {
    "Human": ["Blue", "Brown", "Green", "Hazel", "Amber"],
    "Orc": ["Black", "Brown", "Red"],
    "Dwarf": ["Black", "Brown"],
    "Saurian": ["Purple", "Red", "Black"],
    "Halfling": ["Blue", "Brown", "Green", "Hazel", "Amber"],
    "Aelf": ["Blue", "Brown", "Green", "Hazel", "Amber", "Lilac", "Red"],
    "Faerie": ["Gold", "Red", "Green", "Blue", "Magenta"],
    "Centaur": ["Blue", "Brown", "Green", "Hazel", "Amber"]
}


# Convert inches to feet and inches format like "5'11\""
def convert_inches_to_feet_inches(total_inches):
    feet = total_inches // 12
    inches = total_inches % 12
    height_string = str(feet) + "'" + str(inches) + '"'
    return height_string


# Generate physical traits (height, weight, hair color, eye color) based on race
def generate_physical_traits(race):
    # get height range
    min_height_inches, max_height_inches = RACE_HEIGHT_RANGES[race]
    
    # pick random height
    height_inches = random.randint(min_height_inches, max_height_inches)
    
    # convert to feet and inches
    height_string = convert_inches_to_feet_inches(height_inches)
    
    # get weight range
    min_weight, max_weight = RACE_WEIGHT_RANGES[race]
    
    # pick random weight
    weight = random.randint(min_weight, max_weight)
    
    # get hair colors
    hair_colors = RACE_HAIR_COLORS[race]
    
    # pick hair color
    hair_color = random.choice(hair_colors)
    
    # get eye colors
    eye_colors = RACE_EYE_COLORS[race]
    
    # pick eye color
    eye_color = random.choice(eye_colors)
    
    # return everything
    physical_traits = {
        "height": height_string,
        "weight": weight,
        "hair_color": hair_color,
        "eye_color": eye_color
    }
    
    return physical_traits


# ========================
# PART 4: STATS GENERATOR (DICE MECHANICS)
# ========================

# Roll one die with the given number of sides
def roll_die(sides):
    result = random.randint(1, sides)
    return result


# Roll multiple dice and return a list of results
def roll_dice(number_of_dice, sides):
    dice_results = []
    for i in range(number_of_dice):
        result = roll_die(sides)
        dice_results.append(result)
    return dice_results


# Roll Strength: roll 4d6, drop lowest, sum remaining three
def roll_strength():
    dice_rolls = roll_dice(4, 6)
    dice_rolls.sort()
    dice_rolls.pop(0)  # remove lowest
    total = sum(dice_rolls)
    return total


# Roll Logic: roll 3d6, reroll 1s until all are 2+, then sum
def roll_logic():
    dice_rolls = roll_dice(3, 6)
    
    # reroll 1s
    for i in range(len(dice_rolls)):
        while dice_rolls[i] == 1:
            dice_rolls[i] = roll_die(6)
    
    total = sum(dice_rolls)
    return total


# Roll Agility: roll 5d6, drop highest, sum remaining four, cap at 18
def roll_agility():
    dice_rolls = roll_dice(5, 6)
    dice_rolls.sort()
    dice_rolls.pop()  # remove highest
    total = sum(dice_rolls)
    if total > 18:
        total = 18
    return total


# Roll Intellect: roll 2d6 and add 6
def roll_intellect():
    rolls = roll_dice(2, 6)
    total = sum(rolls) + 6
    return total


# Roll Toughness: roll 4d6, reroll 1s, drop lowest, sum remaining three
def roll_toughness():
    dice_rolls = roll_dice(4, 6)
    
    # reroll 1s
    for i in range(len(dice_rolls)):
        while dice_rolls[i] == 1:
            dice_rolls[i] = roll_die(6)
    
    dice_rolls.sort()
    dice_rolls.pop(0)  # remove lowest
    total = sum(dice_rolls)
    return total


# Roll Charm: roll 1d20, subtract 2, clamp to 3-18 range
def roll_charm():
    result = roll_die(20)
    total = result - 2
    
    if total < 3:
        total = 3
    if total > 18:
        total = 18
    
    return total


# Generate all six stats using the dice rules
def generate_stats():
    strength = roll_strength()
    logic = roll_logic()
    agility = roll_agility()
    intellect = roll_intellect()
    toughness = roll_toughness()
    charm = roll_charm()
    
    stats = {
        "Strength": strength,
        "Logic": logic,
        "Agility": agility,
        "Intellect": intellect,
        "Toughness": toughness,
        "Charm": charm
    }
    
    return stats


# ========================
# PART 5: NAME GENERATOR USING EXCEL
# ========================

# Load name data from Excel file
def load_name_data(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    name_data = {}
    
    # Load Borealan Human names
    if "Borealan Human names" in workbook.sheetnames:
        sheet = workbook["Borealan Human names"]
        male_first = []
        female_first = []
        last_names = []
        
        # Column A has male names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Column E has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=5).value is not None:
            last_name = sheet.cell(row=row, column=5).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["human_borealan"] = {
            "male_first": male_first,
            "female_first": female_first,
            "last": last_names
        }
    
    # Load Tropicanis Human names
    if "Tropicanis Human Names" in workbook.sheetnames:
        sheet = workbook["Tropicanis Human Names"]
        male_first = []
        female_first = []
        last_names = []
        
        # Column A has male names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Column E has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=5).value is not None:
            last_name = sheet.cell(row=row, column=5).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["human_tropicanis"] = {
            "male_first": male_first,
            "female_first": female_first,
            "last": last_names
        }
    
    # Load Selvarossa Human names
    if "Selvarossa Humans" in workbook.sheetnames:
        sheet = workbook["Selvarossa Humans"]
        male_first = []
        female_first = []
        last_names = []
        
        # Column A has male names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Column E has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=5).value is not None:
            last_name = sheet.cell(row=row, column=5).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["human_selvarossa"] = {
            "male_first": male_first,
            "female_first": female_first,
            "last": last_names
        }
    
    # Load Northern Orc names
    if "Northern Orc names" in workbook.sheetnames:
        sheet = workbook["Northern Orc names"]
        first_names = []
        last_names = []
        
        # Column A has androgynous first names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            first_name = sheet.cell(row=row, column=1).value
            first_names.append(str(first_name))
            row += 1
        
        # Column C has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            last_name = sheet.cell(row=row, column=3).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["orc_northern"] = {
            "first": first_names,
            "last": last_names
        }
    
    # Load Dwarven names
    if "Dwarven Names (All regions)" in workbook.sheetnames:
        sheet = workbook["Dwarven Names (All regions)"]
        male_first = []
        female_first = []
        
        # Column A has male names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Dwarves use human last names from Borealan sheet (we'll handle this in generate_name)
        name_data["dwarf"] = {
            "male_first": male_first,
            "female_first": female_first
        }
    
    # Load Aelven names
    if "Aelven names" in workbook.sheetnames:
        sheet = workbook["Aelven names"]
        male_first = []
        female_first = []
        last_names = []
        
        # Column A has male names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Column E has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=5).value is not None:
            last_name = sheet.cell(row=row, column=5).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["aelf"] = {
            "male_first": male_first,
            "female_first": female_first,
            "last": last_names
        }
    
    # Load Centaur names
    if "Centaur names" in workbook.sheetnames:
        sheet = workbook["Centaur names"]
        male_first = []
        female_first = []
        
        # Column A has male names starting at row 4 (row 3 is header)
        row = 4
        while sheet.cell(row=row, column=1).value is not None:
            male_name = sheet.cell(row=row, column=1).value
            male_first.append(str(male_name))
            row += 1
        
        # Column C has female names starting at row 4
        row = 4
        while sheet.cell(row=row, column=3).value is not None:
            female_name = sheet.cell(row=row, column=3).value
            female_first.append(str(female_name))
            row += 1
        
        # Centaurs don't have last names
        name_data["centaur"] = {
            "male_first": male_first,
            "female_first": female_first
        }
    
    # Load Saurian names
    if "Saurian names" in workbook.sheetnames:
        sheet = workbook["Saurian names"]
        first_names = []
        last_names = []
        
        # Column A has androgynous first names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            first_name = sheet.cell(row=row, column=1).value
            first_names.append(str(first_name))
            row += 1
        
        # Column C has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            last_name = sheet.cell(row=row, column=3).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["saurian"] = {
            "first": first_names,
            "last": last_names
        }
    
    # Load Faerie names
    if "Faerie names" in workbook.sheetnames:
        sheet = workbook["Faerie names"]
        first_names = []
        last_names = []
        
        # Column A has androgynous first names starting at row 3
        row = 3
        while sheet.cell(row=row, column=1).value is not None:
            first_name = sheet.cell(row=row, column=1).value
            first_names.append(str(first_name))
            row += 1
        
        # Column C has last names starting at row 3
        row = 3
        while sheet.cell(row=row, column=3).value is not None:
            last_name = sheet.cell(row=row, column=3).value
            last_names.append(str(last_name))
            row += 1
        
        name_data["faerie"] = {
            "first": first_names,
            "last": last_names
        }
    
    workbook.close()
    return name_data


# Generate a name based on region, race, and gender
def generate_name(region, race, gender, name_data):
    first_name = ""
    last_name = ""
    
    # Handle Humans based on region
    if race == "Human":
        if region == "Borealans":
            if gender == "Male":
                first_name = random.choice(name_data["human_borealan"]["male_first"])
            elif gender == "Female":
                first_name = random.choice(name_data["human_borealan"]["female_first"])
            else:  # Nonbinary
                if random.choice([True, False]):
                    first_name = random.choice(name_data["human_borealan"]["male_first"])
                else:
                    first_name = random.choice(name_data["human_borealan"]["female_first"])
            last_name = random.choice(name_data["human_borealan"]["last"])
        
        elif region == "Tropicanis":
            if gender == "Male":
                first_name = random.choice(name_data["human_tropicanis"]["male_first"])
            elif gender == "Female":
                first_name = random.choice(name_data["human_tropicanis"]["female_first"])
            else:  # Nonbinary
                if random.choice([True, False]):
                    first_name = random.choice(name_data["human_tropicanis"]["male_first"])
                else:
                    first_name = random.choice(name_data["human_tropicanis"]["female_first"])
            last_name = random.choice(name_data["human_tropicanis"]["last"])
        
        elif region == "Selvarossa":
            if gender == "Male":
                first_name = random.choice(name_data["human_selvarossa"]["male_first"])
            elif gender == "Female":
                first_name = random.choice(name_data["human_selvarossa"]["female_first"])
            else:  # Nonbinary
                if random.choice([True, False]):
                    first_name = random.choice(name_data["human_selvarossa"]["male_first"])
                else:
                    first_name = random.choice(name_data["human_selvarossa"]["female_first"])
            last_name = random.choice(name_data["human_selvarossa"]["last"])
        
        elif region == "Verdanian":
            # Verdanian uses Borealan names
            if gender == "Male":
                first_name = random.choice(name_data["human_borealan"]["male_first"])
            elif gender == "Female":
                first_name = random.choice(name_data["human_borealan"]["female_first"])
            else:  # Nonbinary
                if random.choice([True, False]):
                    first_name = random.choice(name_data["human_borealan"]["male_first"])
                else:
                    first_name = random.choice(name_data["human_borealan"]["female_first"])
            last_name = random.choice(name_data["human_borealan"]["last"])
    
    elif race == "Orc":
        # orcs use androgynous names
        first_name = random.choice(name_data["orc_northern"]["first"])
        last_name = random.choice(name_data["orc_northern"]["last"])
    
    elif race == "Dwarf":
        if gender == "Male":
            first_name = random.choice(name_data["dwarf"]["male_first"])
        elif gender == "Female":
            first_name = random.choice(name_data["dwarf"]["female_first"])
        else:  # Nonbinary
            if random.choice([True, False]):
                first_name = random.choice(name_data["dwarf"]["male_first"])
            else:
                first_name = random.choice(name_data["dwarf"]["female_first"])
        # dwarves use human last names
        last_name = random.choice(name_data["human_borealan"]["last"])
    
    elif race == "Saurian":
        first_name = random.choice(name_data["saurian"]["first"])
        last_name = random.choice(name_data["saurian"]["last"])
    
    elif race == "Halfling":
        # halflings use tropicanis human names
        if gender == "Male":
            first_name = random.choice(name_data["human_tropicanis"]["male_first"])
        elif gender == "Female":
            first_name = random.choice(name_data["human_tropicanis"]["female_first"])
        else:  # Nonbinary
            if random.choice([True, False]):
                first_name = random.choice(name_data["human_tropicanis"]["male_first"])
            else:
                first_name = random.choice(name_data["human_tropicanis"]["female_first"])
        last_name = random.choice(name_data["human_tropicanis"]["last"])
    
    elif race == "Aelf":
        if gender == "Male":
            first_name = random.choice(name_data["aelf"]["male_first"])
        elif gender == "Female":
            first_name = random.choice(name_data["aelf"]["female_first"])
        else:  # Nonbinary
            if random.choice([True, False]):
                first_name = random.choice(name_data["aelf"]["male_first"])
            else:
                first_name = random.choice(name_data["aelf"]["female_first"])
        last_name = random.choice(name_data["aelf"]["last"])
    
    elif race == "Faerie":
        first_name = random.choice(name_data["faerie"]["first"])
        last_name = random.choice(name_data["faerie"]["last"])
    
    elif race == "Centaur":
        if gender == "Male":
            first_name = random.choice(name_data["centaur"]["male_first"])
        elif gender == "Female":
            first_name = random.choice(name_data["centaur"]["female_first"])
        else:  # Nonbinary
            if random.choice([True, False]):
                first_name = random.choice(name_data["centaur"]["male_first"])
            else:
                first_name = random.choice(name_data["centaur"]["female_first"])
        # centaurs don't have last names
    
    # put name together
    if last_name:
        full_name = first_name + " " + last_name
    else:
        full_name = first_name
    
    return full_name


# ========================
# PART 6: PUTTING IT ALL TOGETHER
# ========================

# Generate a complete NPC character with background, physical traits, stats, and name
def generate_character(name_data):
    background = generate_background()
    physical_traits = generate_physical_traits(background["race"])
    stats = generate_stats()
    name = generate_name(background["region"], background["race"], background["gender"], name_data)
    
    character = {
        "name": name,
        "region": background["region"],
        "city": background["city"],
        "race": background["race"],
        "gender": background["gender"],
        "height": physical_traits["height"],
        "weight": physical_traits["weight"],
        "hair_color": physical_traits["hair_color"],
        "eye_color": physical_traits["eye_color"],
        "stats": stats
    }
    
    return character


# ========================
# PART 7: USER INTERFACE (MAIN PROGRAM)
# ========================

# Print a character profile
def print_character(character):
    print("========================================")
    print("Name: " + character["name"])
    print("Region: " + character["region"])
    print("City of Origin: " + character["city"])
    print("Race: " + character["race"])
    print("Gender: " + character["gender"])
    print("Physical Description:")
    print("  Height: " + character["height"])
    print("  Weight: " + str(character["weight"]) + " lbs")
    print("  Hair Color: " + character["hair_color"])
    print("  Eye Color: " + character["eye_color"])
    print("Stats:")
    print("  Strength: " + str(character["stats"]["Strength"]))
    print("  Logic: " + str(character["stats"]["Logic"]))
    print("  Agility: " + str(character["stats"]["Agility"]))
    print("  Intellect: " + str(character["stats"]["Intellect"]))
    print("  Toughness: " + str(character["stats"]["Toughness"]))
    print("  Charm: " + str(character["stats"]["Charm"]))
    print("========================================")
    print()


# Main function
def main():
    excel_file_path = "Python Character Generator.xlsx"
    name_data = load_name_data(excel_file_path)
    print("This program generates random NPC (non-player character) profiles for a fantasy game.")
    print("Each character will have a name, region, city, race, gender, physical traits, and stats.")
    print()
    
    # ask for input until we get a valid number
    valid_input = False
    number_of_characters = 0
    
    while not valid_input:
        user_input = input("How many characters would you like to generate? ")
        
        # check if input is all digits
        if user_input.isdigit():
            number_of_characters = int(user_input)
            
            # check if it's positive
            if number_of_characters > 0:
                valid_input = True
            else:
                print("Please enter a positive number (greater than 0).")
        else:
            print("Please enter a valid number.")
    
    
    print()
    print("Generating " + str(number_of_characters) + " character(s)...")
    print()
    
    all_characters = []
    for i in range(number_of_characters):
        char = generate_character(name_data)
        all_characters.append(char)
    
    for char in all_characters:
        print_character(char)


main()

